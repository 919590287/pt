#!/usr/bin/env python3
"""Add Nansha 23A/23B and synchronize Nansha route names/timetables.

The Amap geometry is fetched through the existing crawler implementation in
``/Users/a../模型算法/高德爬虫/getGaode.py``.  The authoritative display names
and departures come from the repaired full timetable workbook.  Dry-run is the
default; ``--apply`` makes a complete backup before replacing any live files.
"""

from __future__ import annotations

import argparse
import datetime as dt
import difflib
import importlib.util
import json
import os
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import shapefile
from openpyxl import load_workbook

from apply_real_route_user_updates import (
    decode_timetable,
    timetable_updates,
    write_departure_csv,
    write_geojson,
)
from update_nansha_routes_from_xls import (
    canonical_route_code,
    endpoint_score,
    is_nansha_label,
    route_label_and_endpoints,
)


DEFAULT_SHP = Path(
    "/Volumes/USB DISK/pt_data/广州市/真实数据/"
    "公交线路站点/线路/routes.shp"
)
DEFAULT_XLSX = Path(
    "/Volumes/USB DISK/数据/南沙公交数据/南沙公交时刻表数据/"
    "最新全量线路时刻表（南沙修复）.xlsx"
)
DEFAULT_AMAP = Path("/Users/a../模型算法/高德爬虫/getGaode.py")
DEFAULT_CACHE = Path("/Users/a../模型算法/高德爬虫/amap_cache")

NEW_ROUTES = {
    "23A": {
        "searches": ("南沙23A路", "南沙23A", "南沙23路"),
        "expected_id": "900000217069",
        "xlsx_label": "南沙23A路",
    },
    "23B": {
        "searches": ("南沙23B路", "南沙23B", "南沙23路"),
        "expected_id": "900000217078",
        "xlsx_label": "南沙23B路",
    },
}

GROUP_ALIASES = {
    "31A": "31",
    "31B": "31",
    "31上行": "31",
    "31下行": "31",
    "63A": "63",
    "63B": "63",
    "63上行": "63",
    "63下行": "63",
    "65大站快": "65快",
    "G5普": "G5",
    "W1A": "W1",
    "W1B": "W1",
    "W1上行": "W1",
    "W1下行": "W1",
}

REVIEWED_LINE_ID_LABELS = {
    # Existing clean stop order: 146751 reaches JinYe Primary School first;
    # 151565 reaches Dongjing Village first.
    "900000146751": "九王庙接驳线(黄山鲁东门站--黄山鲁东门站(往金业小学方向))",
    "900000151565": "九王庙接驳线(黄山鲁东门站--黄山鲁东门站(往东井村方向))",
    "900000122274": "南沙K6路(东莞富民商务中心(政务服务中心)站--蕉门公交总站)",
    "900000122275": "南沙K6路(蕉门公交总站--东莞富民商务中心(政务服务中心)站)",
    # 2026-06-14 official adjustment: Shiqiao departures use the east gate;
    # arrivals use the west gate.  The user-requested west-gate display name
    # intentionally omits the parenthetical alias.
    "900000073094": "南沙65路(市桥汽车站东门(番禺人才市场)站--潭洲车站总站)",
    "900000073093": "南沙65路(潭洲车站总站--市桥汽车站西门站)",
    "900000136824": "南沙65路(快)(市桥汽车站东门(番禺人才市场)站--大岗公交总站)",
    "900000136823": "南沙65路(快)(大岗公交总站--市桥汽车站西门站)",
}


def atomic_write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as output:
            json.dump(payload, output, ensure_ascii=False, indent=2)
            output.write("\n")
            output.flush()
            os.fsync(output.fileno())
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def route_group(label: str) -> str:
    code = canonical_route_code(label)
    return GROUP_ALIASES.get(code, code)


def direction_hint(label: str) -> str:
    compact = re.sub(r"\s+", "", label).upper()
    if compact.endswith("上行"):
        return "up"
    if compact.endswith("下行"):
        return "down"
    code = canonical_route_code(label)
    group = route_group(label)
    if group in {"31", "63", "W1"}:
        if code.endswith("A"):
            return "up"
        if code.endswith("B"):
            return "down"
    return ""


def is_target_label(label: str) -> bool:
    compact = re.sub(r"\s+", "", label)
    return (
        is_nansha_label(compact, set())
        or compact.startswith("广州(南沙)深圳机场快线")
        or compact.startswith("广州（南沙）深圳机场快线")
        or compact.startswith("九王庙接驳线")
    )


def timetable_similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    left_set = set(left.split(";"))
    right_set = set(right.split(";"))
    return len(left_set & right_set) / len(left_set | right_set)


def load_timetable_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows: list[dict[str, Any]] = []
        seen_names: set[str] = set()
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            name = str(values[0] or "").strip()
            readable = str(values[1] or "").strip()
            if not name:
                continue
            label, start, end = route_label_and_endpoints(name)
            if not is_target_label(label):
                continue
            if name in seen_names:
                raise ValueError(f"duplicate target name in XLSX row {source_row}: {name}")
            seen_names.add(name)
            updates = timetable_updates(readable)
            rows.append(
                {
                    "source_row": source_row,
                    "name": name,
                    "label": label,
                    "start": start,
                    "end": end,
                    "group": route_group(label),
                    "direction": direction_hint(label),
                    "readable": readable,
                    "updates": {key: str(value) for key, value in updates.items()},
                }
            )
        if not rows:
            raise ValueError("no Nansha timetable rows found")
        return rows
    finally:
        workbook.close()


def import_amap(path: Path, cache_dir: Path):
    spec = importlib.util.spec_from_file_location("existing_get_gaode", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot import Amap crawler: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    module.CACHE_DIR = str(cache_dir)
    (cache_dir / "search").mkdir(parents=True, exist_ok=True)
    (cache_dir / "detail").mkdir(parents=True, exist_ok=True)
    return module


def fetch_new_routes(amap, timetable_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    xlsx_by_label = {row["label"]: row for row in timetable_rows}
    fetched: list[dict[str, Any]] = []
    for code, config in NEW_ROUTES.items():
        expected_id = str(config["expected_id"])
        search_hits: dict[str, dict[str, Any]] = {}
        for query in config["searches"]:
            for hit in amap.search_line_by_name(query):
                line_id = str(hit.get("id") or "")
                if line_id:
                    search_hits[line_id] = hit
            if expected_id in search_hits:
                break
        if expected_id not in search_hits:
            raise ValueError(
                f"Amap search did not return expected {code} ID {expected_id}; "
                f"returned={sorted(search_hits)}"
            )
        detail = amap.get_line_detail(expected_id)
        lines = [
            line for line in detail.get("buslines", [])
            if str(line.get("id") or "") == expected_id
        ]
        if len(lines) != 1:
            raise ValueError(f"Amap detail {expected_id} returned {len(lines)} exact lines")
        line = lines[0]
        coordinates = amap.parse_polyline(line.get("polyline") or "")
        if len(coordinates) < 2:
            raise ValueError(f"Amap route {expected_id} has no usable geometry")
        xlsx_row = xlsx_by_label.get(str(config["xlsx_label"]))
        if xlsx_row is None:
            raise ValueError(f"XLSX is missing {config['xlsx_label']}")
        stops = line.get("busstops") or []
        fetched.append(
            {
                "code": code,
                "line_id": expected_id,
                "amap_name": str(line.get("name") or ""),
                "xlsx": xlsx_row,
                "coordinates": coordinates,
                "company": line.get("company") if isinstance(line.get("company"), str) else "",
                "point_count": len(coordinates),
                "stop_count": len(stops),
                "first_stop": str(stops[0].get("name") or "") if stops else "",
                "last_stop": str(stops[-1].get("name") or "") if stops else "",
                "distance_km": str(line.get("distance") or ""),
                "bbox": [
                    min(point[0] for point in coordinates),
                    min(point[1] for point in coordinates),
                    max(point[0] for point in coordinates),
                    max(point[1] for point in coordinates),
                ],
            }
        )
    return fetched


def record_dict(record: shapefile._Record) -> dict[str, Any]:
    return {str(key): value for key, value in record.as_dict().items()}


def choose_timetable(
    record: dict[str, Any],
    candidates: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    name = str(record.get("name") or "").strip()
    label, start, end = route_label_and_endpoints(name)
    current_timetable = decode_timetable(str(record.get("timetable") or ""))
    hint = direction_hint(label)

    reviewed_name = REVIEWED_LINE_ID_LABELS.get(str(record.get("line_id") or ""))
    if reviewed_name:
        reviewed = [candidate for candidate in candidates if candidate["name"] == reviewed_name]
        if len(reviewed) != 1:
            raise ValueError(
                f"reviewed mapping target missing for {record.get('line_id')}: {reviewed_name}"
            )
        selected = reviewed[0]
        return selected, {
            "method": "reviewed_line_id",
            "endpoint_score": 1.0,
            "timetable_similarity": timetable_similarity(
                current_timetable, selected["readable"]
            ),
            "score": 1.0,
        }

    exact = [candidate for candidate in candidates if candidate["name"] == name]
    if len(exact) == 1:
        selected = exact[0]
        return selected, {
            "method": "exact_name",
            "endpoint_score": 1.0,
            "timetable_similarity": timetable_similarity(
                current_timetable, selected["readable"]
            ),
            "score": 1.0,
        }

    oriented = [
        candidate for candidate in candidates
        if hint and candidate["direction"] == hint
    ]
    pool = oriented or candidates
    scored: list[tuple[float, float, float, float, dict[str, Any]]] = []
    for candidate in pool:
        forward, _reverse = endpoint_score(
            start, end, candidate["start"], candidate["end"]
        )
        time_score = timetable_similarity(current_timetable, candidate["readable"])
        label_score = difflib.SequenceMatcher(
            None,
            re.sub(r"\s+", "", label).upper(),
            re.sub(r"\s+", "", str(candidate["label"])).upper(),
        ).ratio()
        score = 0.68 * forward + 0.27 * time_score + 0.05 * label_score
        scored.append((score, forward, time_score, label_score, candidate))
    scored.sort(key=lambda item: item[0], reverse=True)
    best = scored[0]
    second = scored[1] if len(scored) > 1 else None

    # A single repaired row can intentionally replace two legacy direction
    # records (currently Nansha 10 and W1 branch).  A unique family candidate is
    # therefore accepted even when its repaired endpoint text is loop-style.
    unique_family = len(candidates) == 1
    confident = (
        bool(oriented)
        or unique_family
        # One terminus was renamed on several repaired routes (for example
        # Nansha 65), while the other terminus still fixes the direction.
        # Candidate-gap validation below remains mandatory in that case.
        or best[1] >= 0.49
        or best[2] >= 0.80
    )
    if not confident:
        raise ValueError(
            f"low-confidence timetable match for {record.get('line_id')} {name}: "
            f"best={best[4]['name']} endpoint={best[1]:.3f} time={best[2]:.3f}"
        )
    if (
        second is not None
        and best[0] - second[0] < 0.035
        and best[1] < 0.98
        and best[2] < 0.98
    ):
        raise ValueError(
            f"ambiguous timetable match for {record.get('line_id')} {name}: "
            f"{best[4]['name']}={best[0]:.3f}, {second[4]['name']}={second[0]:.3f}"
        )
    return best[4], {
        "method": "direction_alias" if oriented else (
            "unique_family" if unique_family else "endpoint_timetable"
        ),
        "endpoint_score": round(best[1], 4),
        "timetable_similarity": round(best[2], 4),
        "label_similarity": round(best[3], 4),
        "score": round(best[0], 4),
    }


def prepare_updates(
    reader: shapefile.Reader,
    timetable_rows: list[dict[str, Any]],
    fetched: list[dict[str, Any]],
) -> tuple[
    dict[str, dict[str, Any]],
    list[dict[str, Any]],
    Counter[str],
    set[str],
]:
    rows_by_group: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in timetable_rows:
        rows_by_group[str(row["group"])].append(row)

    existing_ids: set[str] = set()
    updates: dict[str, dict[str, Any]] = {}
    audit: list[dict[str, Any]] = []
    source_use: Counter[str] = Counter()
    for shape_record in reader.iterShapeRecords():
        record = record_dict(shape_record.record)
        line_id = str(record.get("line_id") or "")
        if not line_id or line_id in existing_ids:
            raise ValueError(f"missing or duplicate line_id in SHP: {line_id!r}")
        existing_ids.add(line_id)
        label, _start, _end = route_label_and_endpoints(str(record.get("name") or ""))
        if not is_target_label(label):
            continue
        group = route_group(label)
        candidates = rows_by_group.get(group, [])
        if not candidates:
            raise ValueError(f"no XLSX route family for {line_id}: {record.get('name')}")
        selected, match = choose_timetable(record, candidates)
        source_use[str(selected["name"])] += 1
        after = {"name": selected["name"], **selected["updates"]}
        before = {
            key: str(record.get(key) or "")
            for key in ("name", "first", "last", "dep_count", "first_dep", "last_dep", "timetable")
        }
        changes = {
            key: {"before": before[key], "after": value}
            for key, value in after.items()
            if before[key] != value
        }
        updates[line_id] = after
        audit.append(
            {
                "line_id": line_id,
                "before_name": before["name"],
                "after_name": selected["name"],
                "xlsx_row": selected["source_row"],
                "match": match,
                "changes": changes,
            }
        )

    for item in fetched:
        if str(item["line_id"]) in existing_ids:
            existing_audit = next(
                row for row in audit if str(row["line_id"]) == str(item["line_id"])
            )
            existing_audit["changes"]["geometry"] = {
                "before": "existing geometry",
                "after": f"fresh Amap geometry ({item['point_count']} points)",
            }
            continue
        selected = item["xlsx"]
        source_use[str(selected["name"])] += 1
        audit.append(
            {
                "line_id": item["line_id"],
                "before_name": "",
                "after_name": selected["name"],
                "xlsx_row": selected["source_row"],
                "match": {
                    "method": "new_amap_geometry_exact_xlsx_label",
                    "endpoint_score": 1.0,
                    "timetable_similarity": 1.0,
                    "score": 1.0,
                },
                "changes": {"feature": {"before": "absent", "after": "added"}},
            }
        )

    unused = [row["name"] for row in timetable_rows if source_use[row["name"]] == 0]
    if unused:
        raise ValueError(f"XLSX target rows were not used: {unused}")
    return updates, audit, source_use, existing_ids


def append_record(fields: list[shapefile.Field], item: dict[str, Any]) -> list[Any]:
    updates = item["xlsx"]["updates"]
    values: dict[str, Any] = {
        "line_id": item["line_id"],
        "dir": "0",
        "route_id": item["line_id"],
        "first": updates["first"],
        "last": updates["last"],
        "interval": "",
        "mode": "bus",
        "name": item["xlsx"]["name"],
        "price": "",
        "company": item["company"],
        "dep_count": int(updates["dep_count"]),
        "first_dep": updates["first_dep"],
        "last_dep": updates["last_dep"],
        "timetable": updates["timetable"],
    }
    return [values.get(field.name, "") for field in fields]


def build_updated_shapefile(
    source: Path,
    updates: dict[str, dict[str, Any]],
    fetched: list[dict[str, Any]],
    existing_ids: set[str],
    temp_dir: Path,
) -> Path:
    reader = shapefile.Reader(str(source), encoding="utf-8")
    output = temp_dir / source.name
    writer = shapefile.Writer(str(output), shapeType=reader.shapeType, encoding="utf-8")
    fields = list(reader.fields[1:])
    fetched_by_id = {str(item["line_id"]): item for item in fetched}
    for field in fields:
        writer.field(field.name, field.field_type, field.size, field.decimal)
    field_names = [field.name for field in fields]
    for shape_record in reader.iterShapeRecords():
        values = record_dict(shape_record.record)
        line_id = str(values.get("line_id") or "")
        if line_id in updates:
            values.update(updates[line_id])
        if line_id in fetched_by_id:
            writer.line([fetched_by_id[line_id]["coordinates"]])
        else:
            writer.shape(shape_record.shape)
        writer.record(*[values.get(name, "") for name in field_names])
    for item in fetched:
        if str(item["line_id"]) in existing_ids:
            continue
        writer.line([item["coordinates"]])
        writer.record(*append_record(fields, item))
    writer.close()
    reader.close()
    for suffix in (".prj", ".cpg"):
        companion = source.with_suffix(suffix)
        if companion.exists():
            shutil.copy2(companion, output.with_suffix(suffix))
    return output


def validate_output(
    path: Path,
    expected_count: int,
    timetable_rows: list[dict[str, Any]],
    fetched: list[dict[str, Any]],
) -> dict[str, Any]:
    reader = shapefile.Reader(str(path), encoding="utf-8")
    records = [record_dict(record) for record in reader.iterRecords()]
    shapes = list(reader.iterShapes())
    reader.close()
    if len(records) != expected_count or len(shapes) != expected_count:
        raise ValueError(
            f"output count mismatch: records={len(records)} shapes={len(shapes)} "
            f"expected={expected_count}"
        )
    ids = [str(record.get("line_id") or "") for record in records]
    if len(ids) != len(set(ids)):
        raise ValueError("output contains duplicate line_id")
    by_id = {str(record["line_id"]): record for record in records}
    for item in fetched:
        record = by_id.get(str(item["line_id"]))
        if record is None:
            raise ValueError(f"new route missing after write: {item['line_id']}")
        if str(record.get("name")) != str(item["xlsx"]["name"]):
            raise ValueError(f"new route name mismatch: {item['line_id']}")
    expected_names = {str(row["name"]) for row in timetable_rows}
    output_names = {str(record.get("name") or "") for record in records}
    missing_names = sorted(expected_names - output_names)
    if missing_names:
        raise ValueError(f"repaired names missing from output: {missing_names}")
    blanks = [record["line_id"] for record in records if not str(record.get("timetable") or "")]
    if blanks:
        raise ValueError(f"blank timetables after write: {blanks}")
    return {
        "features": len(records),
        "unique_line_ids": len(set(ids)),
        "unique_names": len(output_names),
        "xlsx_target_names_present": len(expected_names),
        "blank_timetables": len(blanks),
        "new_line_ids": sorted(str(item["line_id"]) for item in fetched),
    }


def create_backup(source: Path, stamp: str) -> Path:
    root = source.parent.parent
    backup = root / f"备份_南沙修复与23AB新增_{stamp}"
    backup.mkdir(parents=True, exist_ok=False)
    for path in source.parent.glob("routes.*"):
        if path.is_file():
            shutil.copy2(path, backup / path.name)
    return backup


def replace_shapefile(source: Path, generated: Path) -> None:
    for suffix in (".shp", ".shx", ".dbf", ".prj", ".cpg"):
        candidate = generated.with_suffix(suffix)
        if not candidate.exists():
            if suffix in {".shp", ".shx", ".dbf"}:
                raise ValueError(f"generated shapefile is missing {suffix}")
            continue
        destination = source.with_suffix(suffix)
        destination_mode: int | None = None
        if destination.exists():
            destination_mode = destination.stat().st_mode
            os.chmod(candidate, destination_mode)
        # The generated file normally lives on the internal system volume,
        # while the live SHP is on a removable drive.  Stage the bytes beside
        # the destination so the final os.replace remains atomic.
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{destination.name}.", suffix=".tmp", dir=destination.parent
        )
        try:
            with os.fdopen(fd, "wb") as output, candidate.open("rb") as input_stream:
                shutil.copyfileobj(input_stream, output)
                output.flush()
                os.fsync(output.fileno())
            if destination_mode is not None:
                os.chmod(temp_name, destination_mode)
            os.replace(temp_name, destination)
        finally:
            if os.path.exists(temp_name):
                os.unlink(temp_name)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--shp", type=Path, default=DEFAULT_SHP)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--amap-script", type=Path, default=DEFAULT_AMAP)
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    for required in (args.shp, args.xlsx, args.amap_script):
        if not required.exists():
            raise FileNotFoundError(required)

    timetable_rows = load_timetable_rows(args.xlsx)
    amap = import_amap(args.amap_script, args.cache_dir)
    fetched = fetch_new_routes(amap, timetable_rows)

    source_reader = shapefile.Reader(str(args.shp), encoding="utf-8")
    source_count = len(source_reader)
    updates, audit, source_use, existing_ids = prepare_updates(
        source_reader, timetable_rows, fetched
    )
    source_reader.close()

    with tempfile.TemporaryDirectory(prefix="nansha_routes_sync_") as temp_name:
        temp_dir = Path(temp_name)
        generated = build_updated_shapefile(
            args.shp, updates, fetched, existing_ids, temp_dir
        )
        added_count = sum(
            str(item["line_id"]) not in existing_ids for item in fetched
        )
        validation = validate_output(
            generated,
            source_count + added_count,
            timetable_rows,
            fetched,
        )

        changed_names = sum(
            item["before_name"] != item["after_name"] and bool(item["before_name"])
            for item in audit
        )
        changed_timetables = sum(
            "timetable" in item["changes"] and bool(item["before_name"])
            for item in audit
        )
        report: dict[str, Any] = {
            "mode": "apply" if args.apply else "dry-run",
            "source_shp": str(args.shp),
            "source_xlsx": str(args.xlsx),
            "amap_script": str(args.amap_script),
            "source_features": source_count,
            "xlsx_target_rows": len(timetable_rows),
            "existing_routes_matched": len(updates),
            "new_routes": [
                {
                    key: item[key]
                    for key in (
                        "code", "line_id", "amap_name", "point_count", "stop_count",
                        "first_stop", "last_stop", "distance_km", "bbox",
                    )
                }
                for item in fetched
            ],
            "amap_routes_added": added_count,
            "amap_routes_refreshed": len(fetched) - added_count,
            "names_changed": changed_names,
            "timetables_changed": changed_timetables,
            "xlsx_rows_reused_for_legacy_features": sorted(
                name for name, count in source_use.items() if count > 1
            ),
            "match_methods": dict(Counter(item["match"]["method"] for item in audit)),
            "validation": validation,
            "route_updates": audit,
        }

        if args.apply:
            stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = create_backup(args.shp, stamp)
            replace_shapefile(args.shp, generated)

            live_validation = validate_output(
                args.shp,
                source_count + added_count,
                timetable_rows,
                fetched,
            )
            reader = shapefile.Reader(str(args.shp), encoding="utf-8")
            departure_rows = [record_dict(record) for record in reader.iterRecords()]
            reader.close()
            departure_csv = args.shp.parent / "routes_departures.csv"
            geojson = args.shp.parent / "routes_with_departures.geojson"
            write_departure_csv(departure_csv, departure_rows)
            geojson_count = write_geojson(args.shp, geojson)
            if geojson_count != live_validation["features"]:
                raise ValueError(
                    f"GeoJSON feature count mismatch: {geojson_count} != "
                    f"{live_validation['features']}"
                )
            report.update(
                {
                    "timestamp": stamp,
                    "backup": str(backup),
                    "live_validation": live_validation,
                    "routes_departures_csv": str(departure_csv),
                    "routes_departures_rows": len(departure_rows),
                    "geojson": str(geojson),
                    "geojson_features": geojson_count,
                }
            )
            report_path = args.report or (
                args.shp.parent.parent / f"南沙线路修复与23AB新增审计_{stamp}.json"
            )
        else:
            report_path = args.report

        if report_path:
            report["report"] = str(report_path)
            atomic_write_json(report_path, report)

    summary_keys = (
        "mode", "source_features", "xlsx_target_rows", "existing_routes_matched",
        "names_changed", "timetables_changed", "match_methods", "validation",
    )
    print(json.dumps({key: report[key] for key in summary_keys}, ensure_ascii=False, indent=2))
    if args.apply:
        print(json.dumps({key: report[key] for key in (
            "backup", "routes_departures_rows", "geojson_features", "report"
        )}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
